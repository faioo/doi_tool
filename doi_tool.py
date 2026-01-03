# -*- coding: utf-8 -*-
"""
DOI 查询工具 - Windows 桌面应用
通过 Crossref API 为 Excel 表格中的论文自动补齐 DOI 号
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import requests
import threading
import os


class DOIToolApp:
    def __init__(self, root):
        self.root = root
        self.root.title("DOI 查询工具")
        self.root.geometry("600x450")
        self.root.resizable(True, True)
        
        # 设置最小窗口尺寸
        self.root.minsize(500, 400)
        
        # 文件路径变量
        self.file_path = tk.StringVar()
        self.file_path.set("未选择文件")
        
        # 处理状态
        self.is_processing = False
        
        # 创建界面
        self.create_widgets()
        
    def create_widgets(self):
        """创建界面组件"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(
            main_frame, 
            text="DOI 查询工具", 
            font=("Microsoft YaHei", 16, "bold")
        )
        title_label.pack(pady=(0, 10))
        
        # 说明文字
        desc_label = ttk.Label(
            main_frame,
            text="选择包含 Title、Journal、Year 列的 Excel 文件，自动查询并补齐 DOI",
            font=("Microsoft YaHei", 9)
        )
        desc_label.pack(pady=(0, 20))
        
        # 文件选择区域
        file_frame = ttk.LabelFrame(main_frame, text="文件选择", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 文件路径显示
        self.file_label = ttk.Label(
            file_frame, 
            textvariable=self.file_path,
            font=("Microsoft YaHei", 9),
            foreground="gray"
        )
        self.file_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # 选择文件按钮
        self.select_btn = ttk.Button(
            file_frame,
            text="选择文件",
            command=self.select_file
        )
        self.select_btn.pack(side=tk.RIGHT, padx=(10, 0))
        
        # 进度区域
        progress_frame = ttk.LabelFrame(main_frame, text="处理进度", padding="10")
        progress_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            variable=self.progress_var,
            maximum=100,
            mode='determinate'
        )
        self.progress_bar.pack(fill=tk.X, pady=(0, 5))
        
        # 进度文字
        self.progress_label = ttk.Label(
            progress_frame,
            text="等待开始...",
            font=("Microsoft YaHei", 9)
        )
        self.progress_label.pack(anchor=tk.W)
        
        # 状态日志区域
        log_frame = ttk.LabelFrame(main_frame, text="处理日志", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # 创建文本框和滚动条
        log_scroll = ttk.Scrollbar(log_frame)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.log_text = tk.Text(
            log_frame,
            height=8,
            font=("Consolas", 9),
            state=tk.DISABLED,
            yscrollcommand=log_scroll.set
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        log_scroll.config(command=self.log_text.yview)
        
        # 按钮区域
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X)
        
        # 开始处理按钮
        self.start_btn = ttk.Button(
            btn_frame,
            text="开始处理",
            command=self.start_process,
            state=tk.DISABLED
        )
        self.start_btn.pack(side=tk.RIGHT)
        
    def select_file(self):
        """打开文件选择对话框"""
        file_path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[
                ("Excel 文件", "*.xlsx"),
                ("Excel 97-2003 文件", "*.xls"),
                ("所有文件", "*.*")
            ]
        )
        
        if file_path:
            self.file_path.set(file_path)
            self.file_label.config(foreground="black")
            self.start_btn.config(state=tk.NORMAL)
            self.log_message(f"已选择文件: {os.path.basename(file_path)}")
            
    def log_message(self, message):
        """添加日志消息"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        
    def update_progress(self, current, total, message=""):
        """更新进度条"""
        progress = (current / total) * 100 if total > 0 else 0
        self.progress_var.set(progress)
        self.progress_label.config(text=f"进度: {current}/{total} ({progress:.1f}%) {message}")
        
    def get_doi(self, title, journal, year):
        """使用 Crossref API 根据标题、期刊和年份查找 DOI"""
        base_url = "https://api.crossref.org/works"
        params = {
            "query.bibliographic": f"{title} {journal} {year}",
            "rows": 1
        }
        
        try:
            response = requests.get(base_url, params=params, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            if data["message"]["items"]:
                return data["message"]["items"][0]["DOI"]
            else:
                return "Not Found"
        except requests.exceptions.Timeout:
            return "Timeout"
        except requests.exceptions.RequestException as e:
            return f"Error: {str(e)[:50]}"
            
    def process_file(self):
        """处理 Excel 文件"""
        try:
            input_path = self.file_path.get()
            
            # 生成输出文件路径
            dir_name = os.path.dirname(input_path)
            file_name = os.path.basename(input_path)
            name, ext = os.path.splitext(file_name)
            output_path = os.path.join(dir_name, f"{name}_with_doi{ext}")
            
            self.log_message(f"正在读取文件...")
            
            # 读取 Excel 文件
            workbook = openpyxl.load_workbook(input_path)
            sheet = workbook.active
            
            # 获取表头
            headers = [cell.value for cell in sheet[1]]
            
            # 查找必要的列
            try:
                title_col = headers.index("Title") + 1
                journal_col = headers.index("Journal") + 1
                year_col = headers.index("Year") + 1
            except ValueError as e:
                self.root.after(0, lambda: messagebox.showerror(
                    "错误", 
                    "Excel 文件必须包含 Title、Journal、Year 列！\n"
                    f"当前表头: {headers}"
                ))
                return
            
            # 添加 DOI 列
            doi_col = len(headers) + 1
            sheet.cell(row=1, column=doi_col, value="DOI")
            
            # 获取总行数（排除表头）
            total_rows = sheet.max_row - 1
            self.log_message(f"共有 {total_rows} 条记录需要处理")
            
            # 遍历每一行
            success_count = 0
            error_count = 0
            
            for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), 1):
                if not self.is_processing:
                    self.log_message("处理已取消")
                    return
                    
                title = sheet.cell(row=i+1, column=title_col).value or ""
                journal = sheet.cell(row=i+1, column=journal_col).value or ""
                year = sheet.cell(row=i+1, column=year_col).value or ""
                
                # 截断标题用于显示
                display_title = title[:40] + "..." if len(str(title)) > 40 else title
                
                self.root.after(0, lambda t=display_title: self.log_message(f"正在查询: {t}"))
                
                # 查询 DOI
                doi = self.get_doi(title, journal, year)
                
                # 写入结果
                sheet.cell(row=i+1, column=doi_col, value=doi)
                
                if doi.startswith("Error") or doi == "Timeout":
                    error_count += 1
                elif doi != "Not Found":
                    success_count += 1
                
                # 更新进度
                self.root.after(0, lambda c=i, t=total_rows: self.update_progress(c, t))
            
            # 保存文件
            self.log_message(f"正在保存文件...")
            workbook.save(output_path)
            workbook.close()
            
            # 完成
            self.root.after(0, lambda: self.on_complete(
                output_path, total_rows, success_count, error_count
            ))
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("错误", f"处理过程中出错:\n{str(e)}"))
            self.root.after(0, self.reset_ui)
            
    def on_complete(self, output_path, total, success, errors):
        """处理完成后的回调"""
        self.is_processing = False
        self.start_btn.config(text="开始处理", state=tk.NORMAL)
        self.select_btn.config(state=tk.NORMAL)
        
        not_found = total - success - errors
        
        self.log_message("=" * 50)
        self.log_message(f"处理完成！")
        self.log_message(f"成功找到 DOI: {success} 条")
        self.log_message(f"未找到 DOI: {not_found} 条")
        self.log_message(f"查询出错: {errors} 条")
        self.log_message(f"结果已保存至: {os.path.basename(output_path)}")
        
        messagebox.showinfo(
            "完成",
            f"处理完成！\n\n"
            f"成功找到 DOI: {success} 条\n"
            f"未找到 DOI: {not_found} 条\n"
            f"查询出错: {errors} 条\n\n"
            f"结果已保存至:\n{output_path}"
        )
        
    def reset_ui(self):
        """重置界面状态"""
        self.is_processing = False
        self.start_btn.config(text="开始处理", state=tk.NORMAL)
        self.select_btn.config(state=tk.NORMAL)
        self.progress_var.set(0)
        self.progress_label.config(text="等待开始...")
        
    def start_process(self):
        """开始处理"""
        if self.is_processing:
            # 取消处理
            self.is_processing = False
            self.start_btn.config(text="开始处理")
            return
            
        file_path = self.file_path.get()
        
        if not file_path or file_path == "未选择文件":
            messagebox.showwarning("警告", "请先选择一个 Excel 文件！")
            return
            
        if not os.path.exists(file_path):
            messagebox.showerror("错误", "文件不存在！")
            return
            
        # 开始处理
        self.is_processing = True
        self.start_btn.config(text="取消处理")
        self.select_btn.config(state=tk.DISABLED)
        self.progress_var.set(0)
        
        self.log_message("=" * 50)
        self.log_message("开始处理...")
        
        # 在后台线程中处理
        thread = threading.Thread(target=self.process_file, daemon=True)
        thread.start()


def main():
    root = tk.Tk()
    
    # 设置 DPI 感知（Windows）
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except:
        pass
    
    app = DOIToolApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()

